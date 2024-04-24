from tkinter import filedialog
import pandas as pd
import numpy as np
import json
from datetime import datetime, timezone, timedelta
from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from random import choice

import connection
import formatacao_objeto
import tabelas

dataInicio = ''
dataFim = ''
produtosAjustados = []

def formatarData(data):
    data_objeto = datetime.strptime(data, '%d/%m/%Y')
    data_formatada = data_objeto.strftime('%Y%m%d')
    return data_formatada

def formatarDataPedido(data):
    milliseconds_since_epoch = data
    seconds_since_epoch = milliseconds_since_epoch / 1000
    date_object = datetime.fromtimestamp(seconds_since_epoch, timezone.utc)
    formatted_date = date_object.strftime('%d/%m/%Y')
    return formatted_date 


def setarData():
    dataInicio = dtInicio.get()
    dtInicioFormatada = formatarData(dataInicio)
    dataFim = dtFim.get()
    dtFimFormatada = formatarData(dataFim)
    if dtInicioFormatada < dtFimFormatada:
        produtosComposicao = connection.getProdutosComposicao(dtInicioFormatada, dtFimFormatada)
        composicaoSemiAcabados = connection.getCompSemiAcabados(dtInicioFormatada, dtFimFormatada)
                
        if len(produtosComposicao) == 0:
            tamanhoLista = 0
            tabelas.criarTabela(secondFrame)
            return tamanhoLista
        else:
            ajustes = connection.getAjustes(dtInicioFormatada, dtFimFormatada)
            estoque = connection.getEstoque()
            produtosQtdAjustada = formatacao_objeto.calcularQtdProducao(produtosComposicao)
            ajustesAplicados = formatacao_objeto.aplicarAjustes(produtosQtdAjustada, ajustes)
            formatacao_objeto.adicionarEstoque(ajustesAplicados, estoque)
            mp_acabados = formatacao_objeto.somarProdutosEvento(ajustesAplicados, incluirLinhaProducao)
            mp_semiAcabados = criarDictSemiAcabados(mp_acabados, composicaoSemiAcabados, estoque)
            produtos = formatacao_objeto.unirListasComposicao(mp_acabados, mp_semiAcabados, incluirLinhaProducao)
            return produtos 
    else:
        tabelas.criarTabela(secondFrame)
        return None

def checarEventosNaLista():
    for evento in tabelas.tabelaSemana.get_children():
        print(tabelas.tabelaSemana.item(evento))

def setarDataPedidosMeioSemana(tipo_requisicao):
    if tipo_requisicao == 'btn':
        dataInicio = dt_inicio_semana.get()
        dtInicioFormatada = formatarData(dataInicio)
        dataFim = dt_fim_semana.get()
        dtFimFormatada = formatarData(dataFim)
    elif tipo_requisicao == 'timer':
        data_atual = datetime.now() - timedelta(days=1)
        dtInicioFormatada = data_atual.strftime('%Y%m%d')
        #dtInicioFormatada = '20240422'
        dataFim = dt_fim_semana.get()
        dtFimFormatada = formatarData(dataFim)

    #checarEventosNaLista()
    global ajustes_meio_semana
    pedidosMeioSemana = connection.getPedidosMeioSemana(dtInicioFormatada, dtFimFormatada)
    semiacabados = connection.getSemiAcabadosMeioSemana(dtInicioFormatada, dtFimFormatada)
    
    if len(pedidosMeioSemana) == 0:
        tamanho_lista = 0
        return tamanho_lista
    else:
        ajustes = connection.getAjustes(dtInicioFormatada, dtFimFormatada)
        estoque = connection.getEstoque()
        produtosQtdAjustada = formatacao_objeto.calcularQtdProducao(pedidosMeioSemana)
        ajustes_meio_semana = formatacao_objeto.aplicarAjustes(produtosQtdAjustada, ajustes)
        formatacao_objeto.adicionarEstoque(ajustes_meio_semana, estoque)
        mp_acabados = formatacao_objeto.somarProdutosEvento(ajustes_meio_semana, incluirLinhaProducao)
        mp_semiAcabados = criarDictSemiAcabados(mp_acabados, semiacabados, estoque)
        produtos = formatacao_objeto.unirListasComposicao(mp_acabados, mp_semiAcabados, incluirLinhaProducao)
        
        return produtos


def recuperarHoraAtual():
    data_hora_atual = datetime.now()
    formato = "%Y-%m-%d_%H-%M-%S"
    data_hora_formatada = data_hora_atual.strftime(formato)
    return data_hora_formatada


def formatarTabela(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    
    if incluirLinhaProducao.get() ==1:
        ws['A1'] = 'ID'
        ws['B1'] = 'Nome do produto'
        ws['C1'] = 'Classificação'
        ws['D1'] = 'Linha'
        ws['E1'] = 'Estoque'
        ws['F1'] = 'Un. Estoque'
        ws['G1'] = 'Qtd. Produção'
        ws['H1'] = 'Unidade'
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FDDA0D", end_color="FDDA0D", fill_type="solid")
        for cell in ws['E'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['F'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['G'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        for cell in ws['H'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        wb.save(caminho)
    else:
        ws['A1'] = 'ID'
        ws['B1'] = 'Nome do produto'
        ws['C1'] = 'Classificação'
        ws['D1'] = 'Estoque'
        ws['E1'] = 'Un. Estoque'
        ws['F1'] = 'Qtd. Produção'
        ws['G1'] = 'Unidade'
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FDDA0D", end_color="FDDA0D", fill_type="solid")
        for cell in ws['D'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['E'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['F'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        for cell in ws['G'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        
        wb.save(caminho)
    

def gerarArquivoExcel(tipoArquivo, listaProdutos):
    root = Tk()
    root.withdraw()

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Arquivos Excel", "*.xlsx")],
                                               title="Salvar Arquivo Excel",
                                               initialfile=f"{tipoArquivo}--{recuperarHoraAtual()}")

    if not file_path:
        print("Operação cancelada pelo usuário.")
        return

    if not file_path.endswith(".xlsx"):
        file_path += ".xlsx"
        
    formatoTabela = pd.DataFrame(listaProdutos)
    formatoTabela.to_excel(file_path, index=False)
    formatarTabela(file_path)
    print(f"Arquivo salvo em: {file_path}")

somaProdutosEventos = []


def filtrarListas(tipoFiltro, listaCompleta):
    if listaCompleta == None:
        return None
    elif listaCompleta == 0:
        return 0
    else:
        listaFiltrada = list(filter(lambda produto:produto['linha'] == tipoFiltro, listaCompleta))
        return listaFiltrada

def separarProdutosEvento(listaProdutos):
    if trazerTodos.get() or filtrarSal.get() or  filtrarDoces.get() or filtrarConfeitaria.get() or filtrarCanapes.get() or filtrarRefeicoes.get() == 1:
        if trazerTodos.get() == 1:       
            gerarArquivoExcel('LISTA_PEDIDOS', listaProdutos)
        if filtrarSal.get() == 1:
            listaSal = filtrarListas('Sal', listaProdutos)
            gerarArquivoExcel('SAL',listaSal)
        if filtrarDoces.get() == 1:
            listaDoces = filtrarListas('Doces', listaProdutos)
            gerarArquivoExcel('DOCES',listaDoces)
        if filtrarRefeicoes.get() == 1:
            listaRefeicoes = filtrarListas('Refeições', listaProdutos)
            gerarArquivoExcel('REFEICOES',listaRefeicoes)
        if filtrarConfeitaria.get() == 1:
            listaConfeitaria = filtrarListas('Confeitaria', listaProdutos)
            gerarArquivoExcel('CONFEITARIA',listaConfeitaria)
        if filtrarCanapes.get() == 1:
            listaCanapes = filtrarListas('Canapés', listaProdutos)
            gerarArquivoExcel('CANAPES',listaCanapes)
    else:
        messagebox.showinfo("Seleção Inválida", "Selecione o tipo de planilha a ser gerado.")
        return None

# def criarTabelaTeste():
#     global tabelaTeste 
#     tabelaTeste = ttk.Treeview(page2, columns = ('Id', 'Cliente', 'Produto', 'Linha', 'Classificação', 'Data Pedido', 'Data Previsão'), show='headings')
#     tabelaTeste.heading('Id', text='Id')
#     tabelaTeste.heading('Cliente', text='Cliente')
#     tabelaTeste.heading('Produto', text='Produto')
#     tabelaTeste.heading('Linha', text='Linha')
#     tabelaTeste.heading('Classificação', text='Classificação')
#     tabelaTeste.heading('Data Pedido', text='Data Pedido')
#     tabelaTeste.heading('Data Previsão', text='Data Previsão')
#     tabelaTeste.grid(row=2, column=0, columnspan=2, padx=(80,0), pady=10, sticky='nsew')

#     tabelaTeste.column('Id', width=80, anchor=CENTER)
#     tabelaTeste.column('Cliente', width=80, anchor=CENTER)
#     tabelaTeste.column('Produto', width=80, anchor=CENTER)
#     tabelaTeste.column('Linha', width=80, anchor=CENTER)
#     tabelaTeste.column('Classificação', width=80, anchor=CENTER)
#     tabelaTeste.column('Data Pedido', width=80, anchor=CENTER)
#     tabelaTeste.column('Data Previsão', width=80, anchor=CENTER)
    
#     data = (1, 'col1', 'col2', 'col3')
    
#     tabelaTeste.insert(parent='', index=0, values=data)


def selecionarOpcao(event):
    todosProdutos = setarData()
    valorSelecionado = combo.get()
    if valorSelecionado == 'Todos os produtos':
        return todosProdutos
    elif valorSelecionado == 'Sal':
        produtosSal = filtrarListas('Sal', todosProdutos)
        return produtosSal
    elif valorSelecionado == 'Doces':
        produtosDoce = filtrarListas('Doces', todosProdutos)
        return produtosDoce
    elif valorSelecionado == 'Confeitaria':
        produtosConfeitaria = filtrarListas('Confeitaria', todosProdutos)
        return produtosConfeitaria
    elif valorSelecionado == 'Refeições':
        produtosRefeicao = filtrarListas('Refeições', todosProdutos)
        return produtosRefeicao
    elif valorSelecionado == 'Canapés':
        produtosCanapes = filtrarListas('Canapés', todosProdutos)
        return produtosCanapes
    
    print(f"Opção selecionada: {valorSelecionado}")

def formatarDataPedido(data):
    milliseconds_since_epoch = data
    seconds_since_epoch = milliseconds_since_epoch / 1000
    date_object = datetime.fromtimestamp(seconds_since_epoch, timezone.utc)
    formatted_date = date_object.strftime('%d/%m/%Y')
    return formatted_date

def verTodosEventos():
    #print('oi')
    indice = tabelas.tabelaSemana.selection()
    if indice:
        produto = tabelas.tabelaSemana.item(indice)['values'][0]
        produtosFiltrados = list(filter(lambda evento:int(evento['idProdutoComposicao']) == int(produto), ajustes_meio_semana))
        for x in produtosFiltrados:
            print(x)
        abrirOutraJanela(produtosFiltrados)

def abrirOutraJanela(produtosFiltrados):
    nova_janela = Toplevel(root)  # Cria uma nova janela
    nova_janela.title("Nova Janela")
    nova_janela.geometry("950x400")
    produto_selecionado = produtosFiltrados[0]['nomeProdutoComposicao']
    # Adicione widgets ou conteúdo à nova janela aqui
    label = Label(nova_janela, text=f'{produto_selecionado}')
    label.grid(padx=20, pady=20)
    
    tabelas.criarTabelaEvento(nova_janela)
    for x in produtosFiltrados:
        cliente = x['nomeEvento']
        produto = x['nomeProdutoAcabado']
        dataPedido = formatarDataPedido(x['dataPedido'])
        dataPrevisao = formatarDataPedido(x['dataPrevisao'])
        qtdEvento = x['qtdProdutoEvento']
        unidade = x['unidadeAcabado']
        data = (cliente, produto, dataPedido, dataPrevisao, qtdEvento, unidade)
        tabelas.tabelaEventos.insert(parent='', index=0, values=data)
    
    
#def mensagemBanco():

def inserirTabelaTeste(tipo_requisicao):
    produtos_meio_semana = setarDataPedidosMeioSemana(tipo_requisicao)
    
    if produtos_meio_semana == 0 or produtos_meio_semana == None:
        mensagem_banco.configure(text='Nenhum evento foi marcado hoje para essa semana')
        #messagebox.showinfo('Sem eventos', 'Nenhum evento nesse período')
    else:
        qtd_eventos_tabela = len(tabelas.tabelaSemana.get_children()) + len(tabelas.tabelaSemana_semi.get_children())
        qtd_eventos_query = len(produtos_meio_semana)
        if qtd_eventos_tabela != qtd_eventos_query:
            mensagem_banco.configure(text='Houve marcação de eventos hoje para essa semana.')
            #messagebox.showinfo('Novos eventos/encomendas', 'Novos pedidos foram feitos hoje para essa semana.')
            tabelas.tabelaSemana.delete(*tabelas.tabelaSemana.get_children())
            tabelas.tabelaSemana_semi.delete(*tabelas.tabelaSemana_semi.get_children())
            for p in produtos_meio_semana:
                id = p['idProdutoComposicao']
                nome = p['nomeProdutoComposicao']
                classificacao = p['classificacao']
                linha = p['linha']
                estoque = p['estoque']
                unidadeEstoque = p['unidadeEstoque']
                totalProducao = p['totalProducao']
                unidade = p['unidade']
                data = (id, nome, classificacao, linha, estoque, unidadeEstoque, totalProducao, unidade)
                if p['produtoAcabado'] == True:
                    tabelas.tabelaSemana.insert(parent='', index=0, values=data)
                else:
                    tabelas.tabelaSemana_semi.insert(parent='', index=0, values=data)
        else:
            #messagebox.showinfo('Sem novos pedidos hoje', 'Nenhum pedido novo por enquanto.')
            return

def formatarListaSemiAcabados(lista, estoque):
    formatacao_objeto.adicionarEstoque(lista, estoque)
    df = pd.DataFrame(lista)
    df['produtoAcabado'] = False
    df['unidade'] = df['unidadeComposicao'].apply(formatacao_objeto.alterarStringUnidade)
    df['nomeProdutoComposicao'] = df['nomeProdutoComposicao']. apply(formatacao_objeto.alterarStringUnidade)
    df['unidadeEstoque'] = df['unidadeEstoque'].apply(formatacao_objeto.alterarStringUnidade)
    df['totalProducao'] = df.apply(formatacao_objeto.converterKg, axis=1)
    df['unidade'] = df['unidade'].apply(formatacao_objeto.mudarUnidade)
    
    if incluirLinhaProducao.get() == 1:
        df = df[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'linha', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
    else:
        df = df[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
    
    result = formatacao_objeto.converterPJson(df)
    return result

def criarDictSemiAcabados(acabados, semiAcabados, estoque):
    dfAcabados = pd.DataFrame(acabados)

    result = dfAcabados.apply(formatacao_objeto.inserirCol_SemiAcabados, semiAcabados=semiAcabados, incluirLinhaProducao=incluirLinhaProducao, axis=1)
    
    resultJson = result.to_json(orient='records')
    dadosDesserializados = json.loads(resultJson)
    
    listaFinal = [p for p in dadosDesserializados if p]
    concatenacao = np.concatenate(listaFinal)
    listaJson = concatenacao.tolist()
    listaFormatada = formatarListaSemiAcabados(listaJson, estoque)
    return listaFormatada

def inserirNaLista():
    if incluirLinhaProducao.get() == 1:
        produtos = selecionarOpcao(Event)
    else:
        produtos = setarData()
    
    if produtos == None:
        messagebox.showinfo('Data inválida', 'Periodo selecionado inválido')
    elif produtos == 0:
        messagebox.showinfo('Lista vazia', 'Não há eventos nesse período de tempo')    
    else:
        produtosOrdenados = sorted(produtos, key=lambda p:p['nomeProdutoComposicao'], reverse=True)
        
        tabelas.table.delete(*tabelas.table.get_children())
        tabelas.tableSemiAcabados.delete(*tabelas.tableSemiAcabados.get_children())
        #['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade']
        for p in produtosOrdenados:
            if incluirLinhaProducao.get() == 1: 
                id = p['idProdutoComposicao']
                nome = p['nomeProdutoComposicao']
                classificacao = p['classificacao']
                linha = p['linha']
                estoque = p['estoque']
                unidadeEstoque = p['unidadeEstoque']
                totalProducao = p['totalProducao']
                unidade = p['unidade']
                data = (id, nome, classificacao, linha, estoque, unidadeEstoque, totalProducao, unidade)
                if p['produtoAcabado'] == True:
                    tabelas.table.insert(parent='', index=0, values=data)
                else:
                    tabelas.tableSemiAcabados.insert(parent='', index=0, values=data)
            else:
                id = p['idProdutoComposicao']
                nome = p['nomeProdutoComposicao']
                classificacao = p['classificacao']
                estoque = p['estoque']
                unidadeEstoque = p['unidadeEstoque']
                totalProducao = p['totalProducao']
                unidade = p['unidade']
                data = (id, nome, classificacao, estoque, unidadeEstoque, totalProducao, unidade)
                if p['produtoAcabado'] == True:
                    tabelas.table.insert(parent='', index=0, values=data)
                else:
                    tabelas.tableSemiAcabados.insert(parent='', index=0, values=data)
            

def gerarPlanilha():
    produtos = setarData()
    if produtos == None:
        messagebox.showinfo('Data inválida', 'Periodo selecionado inválido')
    elif produtos == 0:
        messagebox.showinfo('Lista vazia', 'Não há eventos nesse período de tempo') 
    else:
        if radiobutton_variable.get() == 1:
            print("Gerar planilha acabados")
            composicao_acabados = list(filter(lambda produto:produto['produtoAcabado'] == True, produtos))
            if incluirLinhaProducao.get() == 1:
                separarProdutosEvento(composicao_acabados)
            else:
                gerarArquivoExcel('COMPRAS', composicao_acabados) 
        elif radiobutton_variable.get() == 2:
            print("SEMI ACABADOS")
            composicao_semiacabados = list(filter(lambda produto:produto['produtoAcabado'] == False, produtos))
            if incluirLinhaProducao.get() == 1:
                separarProdutosEvento(composicao_semiacabados)
            else:
                gerarArquivoExcel('COMPRAS', composicao_semiacabados)

def consultarAttBanco():
    #global hora_atual
    hora_atual = datetime.now()
    hora_ultima_checagem.configure(text=f'Momento da última checagem: {str(hora_atual)}')
    inserirTabelaTeste('timer')
    page2.after(10000, consultarAttBanco)


#Tkinter
root = Tk()
root.title("Gerar pedidos de suprimento")

root.geometry("1150x800")

notebook = ttk.Notebook(root)
notebook.pack(fill=BOTH, expand=True)

page1 = Frame(notebook)
notebook.add(page1, text='Página 1')

mainFrame = Frame(page1)
mainFrame.pack(fill=BOTH, expand=1)

canvas = Canvas(mainFrame)
canvas.pack(side=LEFT, fill=BOTH, expand=1)
#canvas.grid(row=0, column=0, sticky=EW)

scrollbar = ttk.Scrollbar(mainFrame, orient=VERTICAL, command=canvas.yview)
scrollbar.pack(side=RIGHT, fill=Y)
#scrollbar.grid(row=0, rowspan=10, column=1, sticky="ns")

canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e:canvas.configure(scrollregion=canvas.bbox("all")))

secondFrame = Frame(canvas)
canvas.create_window((0, 0), window=secondFrame, anchor="nw")

incluirLinhaProducao = IntVar(value=1)
semLinhaProducao = IntVar()
filtrarSal = IntVar(value=0)
filtrarDoces = IntVar(value=0)
filtrarRefeicoes = IntVar(value=0)
filtrarConfeitaria = IntVar(value=0)
filtrarCanapes = IntVar(value=0)
trazerTodos = IntVar(value=1)

explicacao = Label(secondFrame, text="Selecione abaixo o período de tempo para o qual você quer gerar a lista de\n pedidos de suprimento.", font=("Arial", 14))
explicacao.grid(row=0, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

lbl_dtInicio = Label(secondFrame, text="De:", font=("Arial", 14))
lbl_dtInicio.grid(row=1, padx=(0, 190), column=0, sticky="e")

dtInicio = DateEntry(secondFrame, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dtInicio.grid(row=2, column=0, padx=(150, 0), pady=5, sticky="e")

lbl_dtFim = Label(secondFrame, text="Até:", font=("Arial", 14))
lbl_dtFim.grid(row=1, column=1, padx=(50, 0), pady=5, sticky="w")

dtFim = DateEntry(secondFrame, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dtFim.grid(row=2, column=1, padx=(50, 0), pady=5, sticky="w")

c1 = Checkbutton(secondFrame, text='Gerar documento com linha de produção?',variable=incluirLinhaProducao, onvalue=1, offvalue=0, font=("Arial", 14), height=5, width=5, command=tabelas.atualizarTabela)
c1.grid(row=3, columnspan=2, padx=(150, 0), pady=2, sticky="nsew")

opcoes = ['Todos os produtos', 'Sal', 'Doces', 'Confeitaria', 'Refeições', 'Canapés']
opcaoSelecionada = StringVar()
opcaoSelecionada.set('Todos os produtos')
combo = ttk.Combobox(secondFrame, values=opcoes, textvariable=opcaoSelecionada)
combo.grid(row=4, padx=(160, 100), columnspan=2, sticky='nsew')
combo.bind("<<ComboboxSelected>>", selecionarOpcao)

btn_obter_data = Button(secondFrame, text="Mostrar lista", bg='#C0C0C0', font=("Arial", 16), command=inserirNaLista)
btn_obter_data.grid(row=5, column=0, columnspan=2, padx=(80, 0), pady=2, sticky='nsew')

tabela_acabados = Label(secondFrame, text="Composição de produtos acabados", font=("Arial", 14))
tabela_acabados.grid(row=6, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

#row 7 --> Tabela composição acabados

tabela_semiacabados = Label(secondFrame, text="Composição de produtos semi-acabados", font=("Arial", 14))
tabela_semiacabados.grid(row=8, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

#row 9 --> Tabela composição semi-acabados

txtfiltros = Label(secondFrame, text="Selecione os produtos que você deseja filtrar da lista.", font=("Arial", 14))
txtfiltros.grid(row=10, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

c_todos = Checkbutton(secondFrame, text='Todos',variable=trazerTodos, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_todos.grid(row=11, column=0, padx=(0,95), pady=0, sticky='e')

c_sal = Checkbutton(secondFrame, text='Ref',variable=filtrarRefeicoes, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_sal.grid(row=11, column=0, padx=10, pady=0, sticky='e')

c_doces = Checkbutton(secondFrame, text='Doces',variable=filtrarDoces, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_doces.grid(row=11, column=1, padx=(0,0), pady=0, sticky='w')

c_refeicoes = Checkbutton(secondFrame, text='Sal',variable=filtrarSal, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_refeicoes.grid(row=11, column=1, padx=(85,0), pady=0, sticky='w')

c_canapes = Checkbutton(secondFrame, text='Canapés',variable=filtrarCanapes, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=6)
c_canapes.grid(row=12, column=0, sticky='e')

c_confeitaria = Checkbutton(secondFrame, text='Confeitaria',variable=filtrarConfeitaria, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=8)
c_confeitaria.grid(row=12, column=1, padx=10, sticky='w')

txt_tipo_planilha = Label(secondFrame, text="Qual lista de pedido de suprimento você quer gerar?", font=("Arial", 14))
txt_tipo_planilha.grid(row=13, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

radiobutton_variable = IntVar(value=1)
radio_acabados = Radiobutton(secondFrame, text="Composição acabados", font=("Arial", 14), variable = radiobutton_variable, value = 1)
radio_acabados.grid(row = 14, columnspan=2, padx=(150,0), pady=2, sticky="nsew")
radio_semiacabados = Radiobutton(secondFrame, text="Composição semi-acabados", font=("Arial", 14), variable = radiobutton_variable, value = 2)
radio_semiacabados.grid(row = 15, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

btn_obter_data = Button(secondFrame, text="Gerar Planilhas Excel", bg='#C0C0C0', font=("Arial", 16), command=gerarPlanilha)
btn_obter_data.grid(row=16, column=0, columnspan=2, padx=(80, 0), pady=(10, 30), sticky='nsew')


####################################################
#PÁGINA 2
####################################################

page2 = Frame(notebook)
notebook.add(page2,text='Página 2')

lb1 = Label(page2, text='I am page 2')
lb1.grid(pady=20)

hora_ultima_checagem = Label(page2, text='', bg='#C0C0C0', font=("Arial", 16))
hora_ultima_checagem.grid(row=0, column=0)

mensagem_banco = Label(page2, text='', font=("Arial", 16))
mensagem_banco.grid(row=1, column=0)

dt_inicio_semana = Label(page2, text="De:", font=("Arial", 14))
dt_inicio_semana.grid(row=2, padx=(0, 190), column=0, sticky="e")

dt_inicio_semana = DateEntry(page2, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dt_inicio_semana.grid(row=3, column=0, padx=(150, 0), pady=5, sticky="e")

dt_fim_semana = Label(page2, text="Até:", font=("Arial", 14))
dt_fim_semana.grid(row=2, column=1, padx=(50, 0), pady=5, sticky="w")

dt_fim_semana = DateEntry(page2, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dt_fim_semana.grid(row=3, column=1, padx=(50, 0), pady=5, sticky="w")

btn_pedidos_semana = Button(page2, text="Ver pedidos meio semana", bg='#C0C0C0', font=("Arial", 16), command= lambda: inserirTabelaTeste('btn'))
btn_pedidos_semana.grid(row=4)

btn_mostrar_eventos = Button(page2, text="Ver todos os eventos", bg='#C0C0C0', font=("Arial", 16), command=verTodosEventos)
btn_mostrar_eventos.grid(row=7)

tabelas.criarTabelaMeioSemana(page2)
tabelas.criarTabela(secondFrame)

consultarAttBanco()

root.mainloop()





