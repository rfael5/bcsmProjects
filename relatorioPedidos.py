from tkinter import filedialog
import pandas as pd
import numpy as np
import json
from datetime import datetime, timezone

#tkinter
from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry
from tkinter import messagebox
########################################

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from random import choice

import math

dataInicio = ''
dataFim = ''
produtosAjustados = []

#CONEXAO BANCO DE DADOS
###################################################
#QUERY
from query import *
###################################################

def formatarData(data):
    data_objeto = datetime.strptime(data, '%d/%m/%Y')
    data_formatada = data_objeto.strftime('%Y%m%d')
    return data_formatada

def formatarDataPedido(data):
    milliseconds_since_epoch = data
    seconds_since_epoch = milliseconds_since_epoch / 1000
    date_object = datetime.fromtimestamp(seconds_since_epoch, timezone.utc)
    formatted_date = date_object.strftime('%d/%m/%Y')
    return formatted_date 


def setarData():
    dataInicio = dtInicio.get()
    dtInicioFormatada = formatarData(dataInicio)
    dataFim = dtFim.get()
    dtFimFormatada = formatarData(dataFim)
    if dtInicioFormatada < dtFimFormatada:
        produtosComposicao = getProdutosComposicao(dtInicioFormatada, dtFimFormatada)
        composicaoSemiAcabados = getCompSemiAcabados(dtInicioFormatada, dtFimFormatada)
                
        if len(produtosComposicao) == 0:
            tamanhoLista = 0
            criarTabela()
            return tamanhoLista
        else:
            ajustes = getAjustes(dtInicioFormatada, dtFimFormatada)
            estoque = getEstoque()
            produtosQtdAjustada = calcularQtdProducao(produtosComposicao)
            ajustesAplicados =aplicarAjustes(produtosQtdAjustada, ajustes)
            adicionarEstoque(ajustesAplicados, estoque)
            mp_acabados = somarProdutosEvento(ajustesAplicados)
            mp_semiAcabados = criarDictSemiAcabados(mp_acabados, composicaoSemiAcabados, estoque)
            produtos = unirListasComposicao(mp_acabados, mp_semiAcabados)
            return produtos 
    else:
        criarTabela()
        return None

def setarDataPedidosMeioSemana(tipo_requisicao):
    if tipo_requisicao == 'btn':
        dataInicio = dt_inicio_semana.get()
        dtInicioFormatada = formatarData(dataInicio)
        dataFim = dt_fim_semana.get()
        dtFimFormatada = formatarData(dataFim)
    elif tipo_requisicao == 'timer':
        data_atual = datetime.now()
        dtInicioFormatada = data_atual.strftime('%Y%m%d')
        dataFim = dtFim.get()
        dtFimFormatada = formatarData(dataFim)
    
    global ajustes_meio_semana
    pedidosMeioSemana = getPedidosMeioSemana(dtInicioFormatada, dtFimFormatada)
    semiacabados = getSemiAcabadosMeioSemana(dtInicioFormatada, dtFimFormatada)
    
    if len(pedidosMeioSemana) == 0:
        tamanho_lista = 0
        return tamanho_lista
    else:
        ajustes = getAjustes(dtInicioFormatada, dtFimFormatada)
        estoque = getEstoque()
        produtosQtdAjustada = calcularQtdProducao(pedidosMeioSemana)
        ajustes_meio_semana =aplicarAjustes(produtosQtdAjustada, ajustes)
        adicionarEstoque(ajustes_meio_semana, estoque)
        mp_acabados = somarProdutosEvento(ajustes_meio_semana)
        mp_semiAcabados = criarDictSemiAcabados(mp_acabados, semiacabados, estoque)
        produtos = unirListasComposicao(mp_acabados, mp_semiAcabados)
        
        return produtos

    
def filtrarPedidosMeioSemana(dtInicio, dtFim, listaPedidos):
    pedidosMeioSemana = []
    for x in listaPedidos:
            data_pedido = formatarDataPedido(x['dataPedido'])
            data_evento = formatarDataPedido(x['dataEvento'])
            data_previsao = formatarDataPedido(x['dataPrevisao'])
            #teste = formatarDataPedido(dtInicioFormatada)
            if(dataInicio < data_pedido < dataFim and data_previsao < dataFim):
                pedidosMeioSemana.append(x)
    return pedidosMeioSemana

def adicionarAjustes(evento, ajustes):
    for a in ajustes:
        if a['idMovtoped'] == evento['idMovtoped']:
            evento['qtdProdutoEvento'] = evento['qtdProdutoEvento'] + a['ajuste']


def recuperarHoraAtual():
    data_hora_atual = datetime.now()
    formato = "%Y-%m-%d_%H-%M-%S"
    data_hora_formatada = data_hora_atual.strftime(formato)
    return data_hora_formatada

def inserirCol_SemiAcabados(row, semiAcabados):
    listaComposicao = []
    listaOrdenada = sorted(semiAcabados, key=lambda p:p['nomeProdutoAcabado'])
    for p in listaOrdenada:
        if p['idProdutoAcabado'] == row['idProdutoComposicao']:  
            comp_semiacabados = {}  
            comp_semiacabados['idProdutoComposicao'] = p['idProduto']
            comp_semiacabados['nomeProdutoComposicao'] = p['nomeProdutoComposicao']
            comp_semiacabados['qtdComposicao'] = p['qtdProdutoComposicao']
            comp_semiacabados['unidadeComposicao'] = p['unidadeProdutoComposicao']
            comp_semiacabados['classificacao'] = p['classificacao']
            if incluirLinhaProducao.get() == 1:
                comp_semiacabados['linha'] = row['linha']
            comp_semiacabados['idProdutoAcabado'] = p['idProdutoAcabado']
            comp_semiacabados['nomeProdutoAcabado'] = p['nomeProdutoAcabado']
            comp_semiacabados['qtdProducao'] = row['totalProducao']
            comp_semiacabados['unidadeAcabado'] = row['unidade']
            comp_semiacabados['totalProducao'] = (p['qtdProdutoComposicao'] * row['totalProducao']) / p['rendimento']
            if comp_semiacabados != []:
                listaComposicao.append(comp_semiacabados)
    
    return listaComposicao


def converterPJson(lista):
    resultJson = lista.to_json(orient='records')
    dadosDesserializados = json.loads(resultJson)
    return dadosDesserializados

      
def criarDictSemiAcabados(acabados, semiAcabados, estoque):
    dfAcabados = pd.DataFrame(acabados)

    result = dfAcabados.apply(inserirCol_SemiAcabados, semiAcabados=semiAcabados, axis=1)
    
    resultJson = result.to_json(orient='records')
    dadosDesserializados = json.loads(resultJson)
    
    listaFinal = [p for p in dadosDesserializados if p]
    concatenacao = np.concatenate(listaFinal)
    listaJson = concatenacao.tolist()
    listaFormatada = formatarListaSemiAcabados(listaJson, estoque)
    return listaFormatada

def calcTotalProdSemiAcabados(row):
    if 'qtdComposicao' in row and 'qtdProducao' in row:
        totalProducao = int(row['qtdComposicao']) * int(row['qtdProducao'])


def alterarStringUnidade(unidade):
    if '\x00' in unidade:
        unidadeCorrigida = unidade.replace('\x00', '')
        return unidadeCorrigida
    else:
        return unidade   

def converterKg(produto):
    if str(produto['unidade']) == "GR" or str(produto['unidade']) == "ML":
        result = produto['totalProducao'] / 1000 
    else:
        result = produto['totalProducao']
    return round(result, 4)

def formatarListaSemiAcabados(lista, estoque):
    adicionarEstoque(lista, estoque)
    df = pd.DataFrame(lista)
    df['produtoAcabado'] = False
    df['unidade'] = df['unidadeComposicao'].apply(alterarStringUnidade)
    df['nomeProdutoComposicao'] = df['nomeProdutoComposicao']. apply(alterarStringUnidade)
    df['unidadeEstoque'] = df['unidadeEstoque'].apply(alterarStringUnidade)
    df['totalProducao'] = df.apply(converterKg, axis=1)
    df['unidade'] = df['unidade'].apply(mudarUnidade)
    
    if incluirLinhaProducao.get() == 1:
        df = df[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'linha', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
    else:
        df = df[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
    
    result = converterPJson(df)
    return result
    
    
def calcularQtdProducao(produtosComposicao):
    for e in produtosComposicao:
        if e['unidadeAcabado'] == 'PP':
            total = (e["qtdProdutoEvento"] / 10) * e["qtdProdutoComposicao"]
            e["totalProducao"] = total
        elif e['unidadeAcabado'] == 'UD':
            total = (e['qtdProdutoEvento'] / 100) * e['qtdProdutoComposicao']
            e['totalProducao'] = total
        elif e['unidadeAcabado'] == 'UM':
            total = (e['qtdProdutoEvento'] / 10) * e['qtdProdutoComposicao']
            e['totalProducao'] = total
        else:
            total = e["qtdProdutoComposicao"] * e["qtdProdutoEvento"]
            e["totalProducao"] = total
    return produtosComposicao


def aplicarAjustes(produtosComposicao, ajustes):
    for p in produtosComposicao:
        adicionarAjustes(p, ajustes)
    return produtosComposicao


def adicionarEstoque(produtos, estoque):
    for p in produtos:
        p['estoque'] = 0
        p['unidadeEstoque'] = ''
        for e in estoque:
            if p['idProdutoComposicao'] == e['RDX_PRODUTO']:
                p['estoque'] = e['SALDOESTOQUE']
                p['unidadeEstoque'] = e['UN']

def formatarTabela(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    
    if incluirLinhaProducao.get() ==1:
        ws['A1'] = 'ID'
        ws['B1'] = 'Nome do produto'
        ws['C1'] = 'Classificação'
        ws['D1'] = 'Linha'
        ws['E1'] = 'Estoque'
        ws['F1'] = 'Un. Estoque'
        ws['G1'] = 'Qtd. Produção'
        ws['H1'] = 'Unidade'
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FDDA0D", end_color="FDDA0D", fill_type="solid")
        for cell in ws['E'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['F'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['G'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        for cell in ws['H'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        wb.save(caminho)
    else:
        ws['A1'] = 'ID'
        ws['B1'] = 'Nome do produto'
        ws['C1'] = 'Classificação'
        ws['D1'] = 'Estoque'
        ws['E1'] = 'Un. Estoque'
        ws['F1'] = 'Qtd. Produção'
        ws['G1'] = 'Unidade'
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FDDA0D", end_color="FDDA0D", fill_type="solid")
        for cell in ws['D'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['E'][1:]:
            cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
        for cell in ws['F'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        for cell in ws['G'][1:]:
            cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        
        wb.save(caminho)
    

def gerarArquivoExcel(tipoArquivo, listaProdutos):
    root = Tk()
    root.withdraw()

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Arquivos Excel", "*.xlsx")],
                                               title="Salvar Arquivo Excel",
                                               initialfile=f"{tipoArquivo}--{recuperarHoraAtual()}")

    if not file_path:
        print("Operação cancelada pelo usuário.")
        return

    if not file_path.endswith(".xlsx"):
        file_path += ".xlsx"
        
    formatoTabela = pd.DataFrame(listaProdutos)
    formatoTabela.to_excel(file_path, index=False)
    formatarTabela(file_path)
    print(f"Arquivo salvo em: {file_path}")

somaProdutosEventos = []

def mudarUnidade(unidade):
    if unidade == 'GR':
        return 'KG'
    elif unidade == 'ML':
        return 'LT'
    else:
        return unidade


def agruparLinhas(produto):
    if '\x00' in produto['linha']:
        produto['linha'] = produto['linha'].replace('\x00', '')
        
    for x in range(1, 5):
        if produto['linha'] == f'S{x}' or produto['linha'] == 'S6':
            return 'Sal'
    
    for x in range(1, 7):
        if produto['linha'] == f'M-{x}' or produto['linha'] == 'Doce Geral':
            return 'Doces'
    
    for x in range(1, 4):
        if produto['linha'] == f'C-{x}':
            return 'Confeitaria'
    
    if produto['linha'] == 'S5':
        return 'Canapés'
    
    if produto['linha'] == 'S7' or produto['linha'] == 'S8':
        return 'Refeições'           

def unirListasComposicao(acabados, semiAcabados):
    for p in acabados:
        p['produtoAcabado'] = True
    for p in semiAcabados:
        acabados.append(p)
    
    df = pd.DataFrame(acabados)
    if incluirLinhaProducao.get() == 1:
        result = df.groupby(['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'unidade', 'linha', 'estoque', 'unidadeEstoque','produtoAcabado'])[['totalProducao']].sum().reset_index()
        result = result[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'linha', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
        
    else:
        result = df.groupby(['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'unidade', 'estoque', 'unidadeEstoque', 'produtoAcabado'])[['totalProducao']].sum().reset_index()
        result = result[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
    
    res = converterPJson(result)
    dadosOrdenados = sorted(res, key=lambda p:p['nomeProdutoComposicao'])
    return dadosOrdenados   


def somarProdutosEvento(produtosComposicao):
    dfComposicao = pd.DataFrame(produtosComposicao)
    dfComposicao.drop_duplicates(inplace=True)
    dfComposicao['produtoAcabado'] = True
    
    if incluirLinhaProducao.get() == 1:
        dfComposicao['unidade'] = dfComposicao['unidadeComposicao'].apply(alterarStringUnidade)
        dfComposicao['totalProducao'] = dfComposicao.apply(converterKg, axis=1)
        dfComposicao['linha'] = dfComposicao.apply(agruparLinhas, axis=1)
        
        result = dfComposicao.groupby(['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'unidade', 'linha', 'estoque', 'unidadeEstoque', 'produtoAcabado'])[['totalProducao']].sum().reset_index()

        result['unidade'] = result['unidade'].apply(mudarUnidade)
        
        result = result[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'linha', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
        
        resultJson = result.to_json(orient='records')
        dadosDesserializados = json.loads(resultJson)
        dadosOrdenados = sorted(dadosDesserializados, key=lambda p:p['nomeProdutoComposicao'])
        #separarProdutosEvento(dadosDesserializados)
        return dadosOrdenados
    else:
        dfComposicao['unidade'] = dfComposicao['unidadeComposicao'].apply(alterarStringUnidade)
        dfComposicao['totalProducao'] = dfComposicao.apply(converterKg, axis=1)
        
        result = dfComposicao.groupby(['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'unidade', 'estoque', 'unidadeEstoque', 'produtoAcabado'])[['totalProducao']].sum().reset_index()

        result['unidade'] = result['unidade'].apply(mudarUnidade)
        
        result = result[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade', 'produtoAcabado']]
        
        resultJson = result.to_json(orient='records')
        dadosDesserializados = json.loads(resultJson)
        dadosOrdenados = sorted(dadosDesserializados, key=lambda p:p['nomeProdutoComposicao'])
        return dadosOrdenados


def filtrarListas(tipoFiltro, listaCompleta):
    if listaCompleta == None:
        return None
    elif listaCompleta == 0:
        return 0
    else:
        listaFiltrada = list(filter(lambda produto:produto['linha'] == tipoFiltro, listaCompleta))
        return listaFiltrada

def separarProdutosEvento(listaProdutos):
    if trazerTodos.get() or filtrarSal.get() or  filtrarDoces.get() or filtrarConfeitaria.get() or filtrarCanapes.get() or filtrarRefeicoes.get() == 1:
        if trazerTodos.get() == 1:       
            gerarArquivoExcel('LISTA_PEDIDOS', listaProdutos)
        if filtrarSal.get() == 1:
            listaSal = filtrarListas('Sal', listaProdutos)
            gerarArquivoExcel('SAL',listaSal)
        if filtrarDoces.get() == 1:
            listaDoces = filtrarListas('Doces', listaProdutos)
            gerarArquivoExcel('DOCES',listaDoces)
        if filtrarRefeicoes.get() == 1:
            listaRefeicoes = filtrarListas('Refeições', listaProdutos)
            gerarArquivoExcel('REFEICOES',listaRefeicoes)
        if filtrarConfeitaria.get() == 1:
            listaConfeitaria = filtrarListas('Confeitaria', listaProdutos)
            gerarArquivoExcel('CONFEITARIA',listaConfeitaria)
        if filtrarCanapes.get() == 1:
            listaCanapes = filtrarListas('Canapés', listaProdutos)
            gerarArquivoExcel('CANAPES',listaCanapes)
    else:
        messagebox.showinfo("Seleção Inválida", "Selecione o tipo de planilha a ser gerado.")
        return None

def criarTabelaTeste():
    global tabelaTeste 
    tabelaTeste = ttk.Treeview(page2, columns = ('Id', 'Cliente', 'Produto', 'Linha', 'Classificação', 'Data Pedido', 'Data Previsão'), show='headings')
    tabelaTeste.heading('Id', text='Id')
    tabelaTeste.heading('Cliente', text='Cliente')
    tabelaTeste.heading('Produto', text='Produto')
    tabelaTeste.heading('Linha', text='Linha')
    tabelaTeste.heading('Classificação', text='Classificação')
    tabelaTeste.heading('Data Pedido', text='Data Pedido')
    tabelaTeste.heading('Data Previsão', text='Data Previsão')
    tabelaTeste.grid(row=2, column=0, columnspan=2, padx=(80,0), pady=10, sticky='nsew')

    tabelaTeste.column('Id', width=80, anchor=CENTER)
    tabelaTeste.column('Cliente', width=80, anchor=CENTER)
    tabelaTeste.column('Produto', width=80, anchor=CENTER)
    tabelaTeste.column('Linha', width=80, anchor=CENTER)
    tabelaTeste.column('Classificação', width=80, anchor=CENTER)
    tabelaTeste.column('Data Pedido', width=80, anchor=CENTER)
    tabelaTeste.column('Data Previsão', width=80, anchor=CENTER)
    
    data = (1, 'col1', 'col2', 'col3')
    
    tabelaTeste.insert(parent='', index=0, values=data)

def criarTabelaMeioSemana():
    global tabelaSemana
    tabelaSemana = ttk.Treeview(page2, columns = ('ID', 'Produto', 'Classificacao', 'Linha', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
    tabelaSemana.heading('ID', text = 'ID')
    tabelaSemana.heading('Produto', text = 'Produto')
    tabelaSemana.heading('Classificacao', text = 'Classificacao')
    tabelaSemana.heading('Linha', text = 'Linha')
    tabelaSemana.heading('Estoque', text = 'Estoque')
    tabelaSemana.heading('Un. Estoque', text = 'Un. Estoque')
    tabelaSemana.heading('Qtd. Producao', text = 'Qtd. Producao')
    tabelaSemana.heading('Unidade', text = 'Unidade')
    tabelaSemana.grid(row=4, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

    tabelaSemana.column('ID', width=80, anchor=CENTER)
    tabelaSemana.column('Produto', width=300, anchor=CENTER)
    tabelaSemana.column('Classificacao', width=160, anchor=CENTER)
    tabelaSemana.column('Linha', width=100, anchor=CENTER)
    tabelaSemana.column('Estoque', width=80, anchor=CENTER)
    tabelaSemana.column('Un. Estoque', width=80, anchor=CENTER)
    tabelaSemana.column('Qtd. Producao', width=100, anchor=CENTER)
    tabelaSemana.column('Unidade', width=80, anchor=CENTER)
    
    global tabelaSemana_semi
    tabelaSemana_semi = ttk.Treeview(page2, columns = ('ID', 'Produto', 'Classificacao', 'Linha', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
    tabelaSemana_semi.heading('ID', text = 'ID')
    tabelaSemana_semi.heading('Produto', text = 'Produto')
    tabelaSemana_semi.heading('Classificacao', text = 'Classificacao')
    tabelaSemana_semi.heading('Linha', text = 'Linha')
    tabelaSemana_semi.heading('Estoque', text = 'Estoque')
    tabelaSemana_semi.heading('Un. Estoque', text = 'Un. Estoque')
    tabelaSemana_semi.heading('Qtd. Producao', text = 'Qtd. Producao')
    tabelaSemana_semi.heading('Unidade', text = 'Unidade')
    tabelaSemana_semi.grid(row=5, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

    tabelaSemana_semi.column('ID', width=80, anchor=CENTER)
    tabelaSemana_semi.column('Produto', width=300, anchor=CENTER)
    tabelaSemana_semi.column('Classificacao', width=160, anchor=CENTER)
    tabelaSemana_semi.column('Linha', width=100, anchor=CENTER)
    tabelaSemana_semi.column('Estoque', width=80, anchor=CENTER)
    tabelaSemana_semi.column('Un. Estoque', width=80, anchor=CENTER)
    tabelaSemana_semi.column('Qtd. Producao', width=100, anchor=CENTER)
    tabelaSemana_semi.column('Unidade', width=80, anchor=CENTER)

def criarTabela():
    global table 
    table = ttk.Treeview(secondFrame, columns = ('ID', 'Produto', 'Classificacao', 'Linha', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
    table.heading('ID', text = 'ID')
    table.heading('Produto', text = 'Produto')
    table.heading('Classificacao', text = 'Classificacao')
    table.heading('Linha', text = 'Linha')
    table.heading('Estoque', text = 'Estoque')
    table.heading('Un. Estoque', text = 'Un. Estoque')
    table.heading('Qtd. Producao', text = 'Qtd. Producao')
    table.heading('Unidade', text = 'Unidade')
    table.grid(row=7, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

    table.column('ID', width=80, anchor=CENTER)
    table.column('Produto', width=300, anchor=CENTER)
    table.column('Classificacao', width=160, anchor=CENTER)
    table.column('Linha', width=100, anchor=CENTER)
    table.column('Estoque', width=80, anchor=CENTER)
    table.column('Un. Estoque', width=80, anchor=CENTER)
    table.column('Qtd. Producao', width=100, anchor=CENTER)
    table.column('Unidade', width=80, anchor=CENTER)
    
    global tableSemiAcabados 
    tableSemiAcabados = ttk.Treeview(secondFrame, columns = ('ID', 'Produto', 'Classificacao', 'Linha', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
    tableSemiAcabados.heading('ID', text = 'ID')
    tableSemiAcabados.heading('Produto', text = 'Produto')
    tableSemiAcabados.heading('Classificacao', text = 'Classificacao')
    tableSemiAcabados.heading('Linha', text = 'Linha')
    tableSemiAcabados.heading('Estoque', text = 'Estoque')
    tableSemiAcabados.heading('Un. Estoque', text = 'Un. Estoque')
    tableSemiAcabados.heading('Qtd. Producao', text = 'Qtd. Producao')
    tableSemiAcabados.heading('Unidade', text = 'Unidade')
    tableSemiAcabados.grid(row=9, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

    tableSemiAcabados.column('ID', width=80, anchor=CENTER)
    tableSemiAcabados.column('Produto', width=300, anchor=CENTER)
    tableSemiAcabados.column('Classificacao', width=160, anchor=CENTER)
    tableSemiAcabados.column('Linha', width=100, anchor=CENTER)
    tableSemiAcabados.column('Estoque', width=80, anchor=CENTER)
    tableSemiAcabados.column('Un. Estoque', width=80, anchor=CENTER)
    tableSemiAcabados.column('Qtd. Producao', width=100, anchor=CENTER)
    tableSemiAcabados.column('Unidade', width=80, anchor=CENTER)


def atualizarTabela():
    global table
    global tableSemiAcabados
    if incluirLinhaProducao.get() != 1: 
        table = ttk.Treeview(secondFrame, columns = ('ID', 'Produto', 'Classificacao', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
        table.heading('ID', text = 'ID')
        table.heading('Produto', text = 'Produto')
        table.heading('Classificacao', text = 'Classificacao')
        table.heading('Estoque', text = 'Estoque')
        table.heading('Un. Estoque', text = 'Un. Estoque')
        table.heading('Qtd. Producao', text = 'Qtd. Producao')
        table.heading('Unidade', text = 'Unidade')
        table.grid(row=7, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

        table.column('ID', width=80, anchor=CENTER)
        table.column('Produto', width=300, anchor=CENTER)
        table.column('Classificacao', width=160, anchor=CENTER)
        table.column('Estoque', width=80, anchor=CENTER)
        table.column('Un. Estoque', width=80, anchor=CENTER)
        table.column('Qtd. Producao', width=100, anchor=CENTER)
        table.column('Unidade', width=80, anchor=CENTER)
        
        tableSemiAcabados = ttk.Treeview(secondFrame, columns = ('ID', 'Produto', 'Classificacao', 'Estoque', 'Un. Estoque', 'Qtd. Producao', 'Unidade'), show = 'headings')
        tableSemiAcabados.heading('ID', text = 'ID')
        tableSemiAcabados.heading('Produto', text = 'Produto')
        tableSemiAcabados.heading('Classificacao', text = 'Classificacao')
        tableSemiAcabados.heading('Estoque', text = 'Estoque')
        tableSemiAcabados.heading('Un. Estoque', text = 'Un. Estoque')
        tableSemiAcabados.heading('Qtd. Producao', text = 'Qtd. Producao')
        tableSemiAcabados.heading('Unidade', text = 'Unidade')
        tableSemiAcabados.grid(row=9, column=0, columnspan=2, padx=(80, 0), pady=10, sticky="nsew")

        tableSemiAcabados.column('ID', width=80, anchor=CENTER)
        tableSemiAcabados.column('Produto', width=300, anchor=CENTER)
        tableSemiAcabados.column('Classificacao', width=160, anchor=CENTER)
        tableSemiAcabados.column('Estoque', width=80, anchor=CENTER)
        tableSemiAcabados.column('Un. Estoque', width=80, anchor=CENTER)
        tableSemiAcabados.column('Qtd. Producao', width=100, anchor=CENTER)
        tableSemiAcabados.column('Unidade', width=80, anchor=CENTER)
    else:
        criarTabela()

def selecionarOpcao(event):
    todosProdutos = setarData()
    valorSelecionado = combo.get()
    if valorSelecionado == 'Todos os produtos':
        return todosProdutos
    elif valorSelecionado == 'Sal':
        produtosSal = filtrarListas('Sal', todosProdutos)
        return produtosSal
    elif valorSelecionado == 'Doces':
        produtosDoce = filtrarListas('Doces', todosProdutos)
        return produtosDoce
    elif valorSelecionado == 'Confeitaria':
        produtosConfeitaria = filtrarListas('Confeitaria', todosProdutos)
        return produtosConfeitaria
    elif valorSelecionado == 'Refeições':
        produtosRefeicao = filtrarListas('Refeições', todosProdutos)
        return produtosRefeicao
    elif valorSelecionado == 'Canapés':
        produtosCanapes = filtrarListas('Canapés', todosProdutos)
        return produtosCanapes
    
    print(f"Opção selecionada: {valorSelecionado}")

def formatarDataPedido(data):
    milliseconds_since_epoch = data
    seconds_since_epoch = milliseconds_since_epoch / 1000
    date_object = datetime.fromtimestamp(seconds_since_epoch, timezone.utc)
    formatted_date = date_object.strftime('%d/%m/%Y')
    return formatted_date

def verTodosEventos():
    #print('oi')
    indice = tabelaSemana.selection()
    if indice:
        produto = tabelaSemana.item(indice)['values'][0]
        produtosFiltrados = list(filter(lambda evento:int(evento['idProdutoComposicao']) == int(produto), ajustes_meio_semana))
        for x in produtosFiltrados:
            print(x)
        abrirOutraJanela(produtosFiltrados)

def abrirOutraJanela(produtosFiltrados):
    nova_janela = Toplevel(root)  # Cria uma nova janela
    nova_janela.title("Nova Janela")
    nova_janela.geometry("950x400")
    produto_selecionado = produtosFiltrados[0]['nomeProdutoComposicao']
    # Adicione widgets ou conteúdo à nova janela aqui
    label = Label(nova_janela, text=f'{produto_selecionado}')
    label.grid(padx=20, pady=20)
    
    criarTabelaEvento(nova_janela)
    for x in produtosFiltrados:
        cliente = x['nomeEvento']
        produto = x['nomeProdutoAcabado']
        dataPedido = formatarDataPedido(x['dataPedido'])
        dataPrevisao = formatarDataPedido(x['dataPrevisao'])
        qtdEvento = x['qtdProdutoEvento']
        unidade = x['unidadeAcabado']
        data = (cliente, produto, dataPedido, dataPrevisao, qtdEvento, unidade)
        tabelaEventos.insert(parent='', index=0, values=data)
    

def criarTabelaEvento(nova_janela):
    global tabelaEventos
    tabelaEventos = ttk.Treeview(nova_janela, columns = ('Cliente', 'Produto', 'Data pedido', 'Data previsão', 'Qtd Evento', 'Unidade'), show = 'headings')
    tabelaEventos.heading('Cliente', text = 'Cliente')
    tabelaEventos.heading('Produto', text = 'Produto')
    tabelaEventos.heading('Data pedido', text = 'Data pedido')
    tabelaEventos.heading('Data previsão', text = 'Data previsão')
    tabelaEventos.heading('Qtd Evento', text = 'Qtd Evento')
    tabelaEventos.heading('Unidade', text = 'Unidade')
    tabelaEventos.grid(row=1, column=0, padx=(80, 0), pady=10, sticky="nsew")

    tabelaEventos.column('Cliente', width=160, anchor=CENTER)
    tabelaEventos.column('Produto', width=320, anchor=CENTER)
    tabelaEventos.column('Data pedido', width=80, anchor=CENTER)
    tabelaEventos.column('Data previsão', width=80, anchor=CENTER)
    tabelaEventos.column('Qtd Evento', width=80, anchor=CENTER)
    tabelaEventos.column('Unidade', width=60, anchor=CENTER)
    
#def mensagemBanco():

def inserirTabelaTeste(tipo_requisicao):
    produtos_meio_semana = setarDataPedidosMeioSemana(tipo_requisicao)
    
    if produtos_meio_semana == 0 or produtos_meio_semana == None:
        mensagem_banco.configure(text='Nenhum evento foi marcado hoje para essa semana')
    else:
        tabelaSemana.delete(*tabelaSemana.get_children())
        tabelaSemana_semi.delete(*tabelaSemana_semi.get_children())
        for p in produtos_meio_semana:
            id = p['idProdutoComposicao']
            nome = p['nomeProdutoComposicao']
            classificacao = p['classificacao']
            linha = p['linha']
            estoque = p['estoque']
            unidadeEstoque = p['unidadeEstoque']
            totalProducao = p['totalProducao']
            unidade = p['unidade']
            data = (id, nome, classificacao, linha, estoque, unidadeEstoque, totalProducao, unidade)
            if p['produtoAcabado'] == True:
                tabelaSemana.insert(parent='', index=0, values=data)
            else:
                tabelaSemana_semi.insert(parent='', index=0, values=data) 
         

def inserirNaLista():
    if incluirLinhaProducao.get() == 1:
        produtos = selecionarOpcao(Event)
    else:
        produtos = setarData()
    
    if produtos == None:
        messagebox.showinfo('Data inválida', 'Periodo selecionado inválido')
    elif produtos == 0:
        messagebox.showinfo('Lista vazia', 'Não há eventos nesse período de tempo')    
    else:
        produtosOrdenados = sorted(produtos, key=lambda p:p['nomeProdutoComposicao'], reverse=True)
        
        table.delete(*table.get_children())
        tableSemiAcabados.delete(*tableSemiAcabados.get_children())
        #['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade']
        for p in produtosOrdenados:
            if incluirLinhaProducao.get() == 1: 
                id = p['idProdutoComposicao']
                nome = p['nomeProdutoComposicao']
                classificacao = p['classificacao']
                linha = p['linha']
                estoque = p['estoque']
                unidadeEstoque = p['unidadeEstoque']
                totalProducao = math.ceil(p['totalProducao'])
                unidade = p['unidade']
                data = (id, nome, classificacao, linha, estoque, unidadeEstoque, totalProducao, unidade)
                if p['produtoAcabado'] == True:
                    table.insert(parent='', index=0, values=data)
                else:
                    tableSemiAcabados.insert(parent='', index=0, values=data)
            else:
                id = p['idProdutoComposicao']
                nome = p['nomeProdutoComposicao']
                classificacao = p['classificacao']
                estoque = p['estoque']
                unidadeEstoque = p['unidadeEstoque']
                totalProducao = p['totalProducao']
                unidade = p['unidade']
                data = (id, nome, classificacao, estoque, unidadeEstoque, totalProducao, unidade)
                if p['produtoAcabado'] == True:
                    table.insert(parent='', index=0, values=data)
                else:
                    tableSemiAcabados.insert(parent='', index=0, values=data)
            

def gerarPlanilha():
    produtos = setarData()
    if produtos == None:
        messagebox.showinfo('Data inválida', 'Periodo selecionado inválido')
    elif produtos == 0:
        messagebox.showinfo('Lista vazia', 'Não há eventos nesse período de tempo') 
    else:

        if radiobutton_variable.get() == 1:
            print("Gerar planilha acabados")
            composicao_acabados = list(filter(lambda produto:produto['produtoAcabado'] == True, produtos))
            if incluirLinhaProducao.get() == 1:
                separarProdutosEvento(composicao_acabados)
            else:
                gerarArquivoExcel('COMPRAS', composicao_acabados) 
        elif radiobutton_variable.get() == 2:
            print("SEMI ACABADOS")
            composicao_semiacabados = list(filter(lambda produto:produto['produtoAcabado'] == False, produtos))
            if incluirLinhaProducao.get() == 1:
                separarProdutosEvento(composicao_semiacabados)
            else:
                gerarArquivoExcel('COMPRAS', composicao_semiacabados)

def consultarAttBanco():
    #global hora_atual
    hora_atual = datetime.now()
    hora_ultima_checagem.configure(text=f'Momento da última checagem: {str(hora_atual)}')
    inserirTabelaTeste('timer')
    page2.after(10000, consultarAttBanco)


#Tkinter
root = Tk()
root.title("Gerar pedidos de suprimento")

root.geometry("1150x800")

notebook = ttk.Notebook(root)
notebook.pack(fill=BOTH, expand=True)

page1 = Frame(notebook)
notebook.add(page1, text='Página 1')

mainFrame = Frame(page1)
mainFrame.pack(fill=BOTH, expand=1)

canvas = Canvas(mainFrame)
canvas.pack(side=LEFT, fill=BOTH, expand=1)
#canvas.grid(row=0, column=0, sticky=EW)

scrollbar = ttk.Scrollbar(mainFrame, orient=VERTICAL, command=canvas.yview)
scrollbar.pack(side=RIGHT, fill=Y)
#scrollbar.grid(row=0, rowspan=10, column=1, sticky="ns")

canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e:canvas.configure(scrollregion=canvas.bbox("all")))

secondFrame = Frame(canvas)
canvas.create_window((0, 0), window=secondFrame, anchor="nw")

incluirLinhaProducao = IntVar(value=1)
semLinhaProducao = IntVar()
filtrarSal = IntVar(value=0)
filtrarDoces = IntVar(value=0)
filtrarRefeicoes = IntVar(value=0)
filtrarConfeitaria = IntVar(value=0)
filtrarCanapes = IntVar(value=0)
trazerTodos = IntVar(value=1)

explicacao = Label(secondFrame, text="Selecione abaixo o período de tempo para o qual você quer gerar a lista de\n pedidos de suprimento.", font=("Arial", 14))
explicacao.grid(row=0, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

lbl_dtInicio = Label(secondFrame, text="De:", font=("Arial", 14))
lbl_dtInicio.grid(row=1, padx=(0, 190), column=0, sticky="e")

dtInicio = DateEntry(secondFrame, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dtInicio.grid(row=2, column=0, padx=(150, 0), pady=5, sticky="e")

lbl_dtFim = Label(secondFrame, text="Até:", font=("Arial", 14))
lbl_dtFim.grid(row=1, column=1, padx=(50, 0), pady=5, sticky="w")

dtFim = DateEntry(secondFrame, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dtFim.grid(row=2, column=1, padx=(50, 0), pady=5, sticky="w")

c1 = Checkbutton(secondFrame, text='Gerar documento com linha de produção?',variable=incluirLinhaProducao, onvalue=1, offvalue=0, font=("Arial", 14), height=5, width=5, command=atualizarTabela)
c1.grid(row=3, columnspan=2, padx=(150, 0), pady=2, sticky="nsew")

opcoes = ['Todos os produtos', 'Sal', 'Doces', 'Confeitaria', 'Refeições', 'Canapés']
opcaoSelecionada = StringVar()
opcaoSelecionada.set('Todos os produtos')
combo = ttk.Combobox(secondFrame, values=opcoes, textvariable=opcaoSelecionada)
combo.grid(row=4, padx=(160, 100), columnspan=2, sticky='nsew')
combo.bind("<<ComboboxSelected>>", selecionarOpcao)

btn_obter_data = Button(secondFrame, text="Mostrar lista", bg='#C0C0C0', font=("Arial", 16), command=inserirNaLista)
btn_obter_data.grid(row=5, column=0, columnspan=2, padx=(80, 0), pady=2, sticky='nsew')

tabela_acabados = Label(secondFrame, text="Composição de produtos acabados", font=("Arial", 14))
tabela_acabados.grid(row=6, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

#row 7 --> Tabela composição acabados

tabela_semiacabados = Label(secondFrame, text="Composição de produtos semi-acabados", font=("Arial", 14))
tabela_semiacabados.grid(row=8, columnspan=2, padx=(150, 0), pady=10, sticky="nsew")

#row 9 --> Tabela composição semi-acabados

txtfiltros = Label(secondFrame, text="Selecione os produtos que você deseja filtrar da lista.", font=("Arial", 14))
txtfiltros.grid(row=10, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

c_todos = Checkbutton(secondFrame, text='Todos',variable=trazerTodos, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_todos.grid(row=11, column=0, padx=(0,95), pady=0, sticky='e')

c_sal = Checkbutton(secondFrame, text='Ref',variable=filtrarRefeicoes, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_sal.grid(row=11, column=0, padx=10, pady=0, sticky='e')

c_doces = Checkbutton(secondFrame, text='Doces',variable=filtrarDoces, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_doces.grid(row=11, column=1, padx=(0,0), pady=0, sticky='w')

c_refeicoes = Checkbutton(secondFrame, text='Sal',variable=filtrarSal, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=5)
c_refeicoes.grid(row=11, column=1, padx=(85,0), pady=0, sticky='w')

c_canapes = Checkbutton(secondFrame, text='Canapés',variable=filtrarCanapes, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=6)
c_canapes.grid(row=12, column=0, sticky='e')

c_confeitaria = Checkbutton(secondFrame, text='Confeitaria',variable=filtrarConfeitaria, onvalue=1, offvalue=0, font=("Arial", 14), height=2, width=8)
c_confeitaria.grid(row=12, column=1, padx=10, sticky='w')

txt_tipo_planilha = Label(secondFrame, text="Qual lista de pedido de suprimento você quer gerar?", font=("Arial", 14))
txt_tipo_planilha.grid(row=13, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

radiobutton_variable = IntVar(value=1)
radio_acabados = Radiobutton(secondFrame, text="Composição acabados", font=("Arial", 14), variable = radiobutton_variable, value = 1)
radio_acabados.grid(row = 14, columnspan=2, padx=(150,0), pady=2, sticky="nsew")
radio_semiacabados = Radiobutton(secondFrame, text="Composição semi-acabados", font=("Arial", 14), variable = radiobutton_variable, value = 2)
radio_semiacabados.grid(row = 15, columnspan=2, padx=(150,0), pady=2, sticky="nsew")

btn_obter_data = Button(secondFrame, text="Gerar Planilhas Excel", bg='#C0C0C0', font=("Arial", 16), command=gerarPlanilha)
btn_obter_data.grid(row=16, column=0, columnspan=2, padx=(80, 0), pady=(10, 30), sticky='nsew')


####################################################
#PÁGINA 2
####################################################

page2 = Frame(notebook)
notebook.add(page2,text='Página 2')

lb1 = Label(page2, text='I am page 2')
lb1.grid(pady=20)

hora_ultima_checagem = Label(page2, text='', bg='#C0C0C0', font=("Arial", 16))
hora_ultima_checagem.grid(row=0, column=0)

mensagem_banco = Label(page2, text='', font=("Arial", 16))
mensagem_banco.grid(row=1, column=0)

dt_inicio_semana = Label(page2, text="De:", font=("Arial", 14))
dt_inicio_semana.grid(row=2, padx=(0, 190), column=0, sticky="e")

dt_inicio_semana = DateEntry(page2, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dt_inicio_semana.grid(row=3, column=0, padx=(150, 0), pady=5, sticky="e")

dt_fim_semana = Label(page2, text="Até:", font=("Arial", 14))
dt_fim_semana.grid(row=2, column=1, padx=(50, 0), pady=5, sticky="w")

dt_fim_semana = DateEntry(page2, font=('Arial', 12), width=22, height=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
dt_fim_semana.grid(row=3, column=1, padx=(50, 0), pady=5, sticky="w")

btn_pedidos_semana = Button(page2, text="Ver pedidos meio semana", bg='#C0C0C0', font=("Arial", 16), command= lambda: inserirTabelaTeste('btn'))
btn_pedidos_semana.grid(row=4)

btn_mostrar_eventos = Button(page2, text="Ver todos os eventos", bg='#C0C0C0', font=("Arial", 16), command=verTodosEventos)
btn_mostrar_eventos.grid(row=7)

criarTabelaMeioSemana()
criarTabela()

consultarAttBanco()

root.mainloop()





